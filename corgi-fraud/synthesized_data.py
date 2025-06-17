import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
from faker import Faker
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill
import json
from typing import Dict, List, Optional
from autogen import AssistantAgent, UserProxyAgent, config_list_from_json
fake = Faker()
import os

class FraudDetectionConfig:
    """Configuration for advanced fraud detection system"""
    LLM_CONFIG = {
        "config_list": [
            {
                "model": "gpt-4",
                "api_key": os.getenv("OPENAI_API_KEY")  # Get key from environment
            }
        ],
        "timeout": 120,
        "temperature": 0.3,
        "seed": 42
    }
    
    GPT4_CONFIG = {
        "config_list": [
            {
                "model": "gpt-4",
                "api_key": os.getenv("OPENAI_API_KEY")
            }
        ],
        "timeout": 180,
        "temperature": 0.2
    }
# Enhanced fraud patterns with agent-specific triggers
FRAUD_PATTERNS = {
    'social_media_fraud': {
        'description': 'Social media inconsistency',
        'text_keywords': ['vacation photos', 'recent activity', 'contradictory posts'],
        'amount_range': (5000, 25000),
        'social_media_flags': ['activity_during_disability', 'location_mismatch', 'deleted_posts'],
        'network_flags': ['new_connections', 'suspicious_friends'],
        'blockchain_flags': [],
        'medical_flags': [],
        'geospatial_flags': []
    },
    'money_laundering': {
        'description': 'Potential money laundering',
        'text_keywords': ['cash transactions', 'unexplained wealth', 'foreign accounts'],
        'amount_range': (15000, 100000),
        'social_media_flags': ['luxury_display', 'sudden_wealth'],
        'network_flags': ['known_fraudsters', 'shell_companies'],
        'blockchain_flags': ['mixing_service', 'darknet_market'],
        'medical_flags': [],
        'geospatial_flags': ['location_spoofing']
    },
    'crypto_theft': {
        'description': 'Fake crypto theft',
        'text_keywords': ['hacked wallet', 'lost keys', 'phishing attack'],
        'amount_range': (10000, 50000),
        'social_media_flags': ['security_negligence', 'password_sharing'],
        'network_flags': ['suspicious_transfers'],
        'blockchain_flags': ['wallet_active', 'funds_moved'],
        'medical_flags': [],
        'geospatial_flags': ['ip_location_mismatch']
    },
    'medical_fraud': {
        'description': 'Healthcare provider fraud',
        'text_keywords': ['unnecessary treatment', 'upcoded procedures', 'phantom billing'],
        'amount_range': (8000, 75000),
        'social_media_flags': ['provider_reputation'],
        'network_flags': ['billing_network'],
        'blockchain_flags': [],
        'medical_flags': ['unusual_frequency', 'improbable_diagnoses'],
        'geospatial_flags': ['impossible_travel']
    }
}

class AdvancedFraudAgents:
    def __init__(self):
        self.agents = self._initialize_agents()
        self.user_proxy = UserProxyAgent(
            name="Admin",
            human_input_mode="NEVER",
            max_consecutive_auto_reply=10,
            code_execution_config=False
        )
    
    def _initialize_agents(self) -> Dict[str, AssistantAgent]:
        """Initialize the complete set of advanced analysis agents"""
        agents = {
            "social": AssistantAgent(
                name="SocialAnalystPro",
                system_message="""Analyze social media for fraud indicators:
                - Post/claim timeline analysis
                - Geolocation verification
                - Sentiment inconsistency
                - Image metadata examination
                - Profile authenticity scoring
                - Network graph analysis
                
                Output JSON format:
                {
                    "score": 0-1, 
                    "confidence": 0-1,
                    "timeline_analysis": {
                        "inconsistencies": [...],
                        "time_gaps": [...]
                    },
                    "geolocation_verification": {
                        "verified": bool,
                        "conflicting_locations": [...]
                    },
                    "media_analysis": {
                        "image_metadata": {...},
                        "deepfake_indicators": float
                    },
                    "recommendations": [...]
                }""",
                llm_config=FraudDetectionConfig.LLM_CONFIG
            ),
            "network": AssistantAgent(
                name="NetworkThreatIntel",
                system_message="""Examine network patterns for anomalies:
                - IP reputation analysis
                - Device fingerprinting
                - Behavioral biometrics
                - Temporal patterns
                - Connection graph analysis
                - Threat intelligence correlation
                
                Output JSON format:
                {
                    "score": 0-1,
                    "threat_level": "low/medium/high/critical",
                    "suspicious_ips": [...],
                    "behavioral_anomalies": {
                        "typing_pattern": float,
                        "mouse_movements": float,
                        "session_timing": float
                    },
                    "network_graph": {
                        "centrality": float,
                        "cluster_coefficient": float
                    },
                    "tor_usage": bool
                }""",
                llm_config=FraudDetectionConfig.LLM_CONFIG
            ),
            "blockchain": AssistantAgent(
                name="BlockchainForensics",
                system_message="""Analyze blockchain transactions:
                - Wallet transaction graph
                - Smart contract interactions
                - Mixer/tumbler detection
                - Dark web affiliations
                - Token flow analysis
                - Anonymity set evaluation
                
                Output JSON format:
                {
                    "score": 0-1,
                    "wallet_risk": {
                        "score": 0-1,
                        "reasons": [...]
                    },
                    "transaction_analysis": {
                        "suspicious_count": int,
                        "mixer_usage": bool,
                        "darknet_affiliation": float
                    },
                    "token_flow": {
                        "source": str,
                        "destination": str,
                        "path": [...]
                    },
                    "recommendations": [...]
                }""",
                llm_config=FraudDetectionConfig.LLM_CONFIG
            ),
            "medical": AssistantAgent(
                name="MedicalClaimsExpert",
                system_message="""Evaluate medical claims:
                - Procedure-code consistency
                - Treatment duration analysis
                - Provider reputation
                - Historical claim patterns
                - Clinical plausibility
                - Billing code analysis
                
                Output JSON format:
                {
                    "score": 0-1,
                    "procedure_analysis": {
                        "consistency": 0-1,
                        "unusual_codes": [...]
                    },
                    "treatment_analysis": {
                        "duration_anomalies": [...],
                        "frequency_issues": [...]
                    },
                    "provider_analysis": {
                        "reputation_score": 0-1,
                        "sanction_history": [...]
                    },
                    "clinical_plausibility": 0-1
                }""",
                llm_config=FraudDetectionConfig.LLM_CONFIG
            ),
            "geospatial": AssistantAgent(
                name="GeospatialAnalyst",
                system_message="""Verify geospatial claims:
                - Location verification
                - Weather correlation
                - Traffic/accident data
                - Satellite imagery analysis
                - IP geolocation
                - Travel feasibility
                
                Output JSON format:
                {
                    "score": 0-1,
                    "location_verification": {
                        "verified": bool,
                        "confidence": 0-1,
                        "sources": [...]
                    },
                    "weather_analysis": {
                        "consistent": bool,
                        "anomalies": [...]
                    },
                    "travel_analysis": {
                        "feasible": bool,
                        "time_required": float,
                        "route_verification": float
                    },
                    "ip_analysis": {
                        "vpn_proxy": bool,
                        "location_mismatch": bool
                    }
                }""",
                llm_config=FraudDetectionConfig.LLM_CONFIG
            ),
            "decision": AssistantAgent(
                name="DecisionEnginePro",
                system_message="""Make final claim determination:
                1. Aggregate agent findings with weighted scoring
                2. Apply business rules and regulatory requirements
                3. Evaluate fraud probability using ensemble methods
                4. Generate audit trail and documentation
                
                Final JSON output format:
                {
                    "decision": "APPROVE/REJECT/ESCALATE",
                    "confidence": 0-1,
                    "risk_factors": [
                        {
                            "factor": str,
                            "score": 0-1,
                            "weight": 0-1
                        }
                    ],
                    "recommendations": {
                        "immediate_actions": [...],
                        "long_term_mitigations": [...]
                    },
                    "audit_trail": {
                        "agent_reports": [...],
                        "timestamps": [...],
                        "evidence_references": [...]
                    },
                    "regulatory_compliance": {
                        "requirements_met": [...],
                        "potential_violations": [...]
                    }
                }""",
                llm_config=FraudDetectionConfig.GPT4_CONFIG
            )
        }
        return agents

    def analyze_claim(self, claim_data: dict) -> dict:
        """Orchestrate multi-agent fraud analysis"""
        analysis_results = {}
        
        # Parallel agent analysis
        analysis_results["social"] = self._analyze_with_agent(
            "social", claim_data.get("social_media", {})
        )
        analysis_results["network"] = self._analyze_with_agent(
            "network", claim_data.get("network_analysis", {})
        )
        analysis_results["blockchain"] = self._analyze_with_agent(
            "blockchain", claim_data.get("blockchain", {})
        )
        analysis_results["medical"] = self._analyze_with_agent(
            "medical", claim_data.get("medical_data", {})
        )
        analysis_results["geospatial"] = self._analyze_with_agent(
            "geospatial", claim_data.get("location_data", {})
        )
        
        # Decision engine processing
        decision_input = {
            "claim_details": claim_data,
            "agent_findings": analysis_results
        }
        final_decision = self._analyze_with_agent("decision", decision_input)
        
        return {
            "analysis_results": analysis_results,
            "final_decision": final_decision
        }

    def _analyze_with_agent(self, agent_name: str, data: dict) -> dict:
        """Execute analysis with a specific agent"""
        try:
            self.user_proxy.initiate_chat(
                self.agents[agent_name],
                message=json.dumps(data, indent=2))
            
            last_message = self.user_proxy.chat_messages[self.agents[agent_name]][-1]
            return json.loads(last_message["content"])
        except Exception as e:
            print(f"Error in {agent_name} analysis: {str(e)}")
            return {"error": str(e)}

def generate_claim_metadata(is_fraud: bool, pattern: dict = None) -> dict:
    """Generate metadata with agent-specific detection triggers"""
    metadata = {
        'claimant_age': random.randint(18, 80),
        'policy_age_days': random.randint(30, 365*5),
        'previous_claims': random.randint(0, 5),
        'credit_score': random.randint(300, 850),
        'digital_footprint': {
            'device_types': random.sample(['mobile', 'desktop', 'tablet'], random.randint(1, 3)),
            'os_versions': [fake.user_agent() for _ in range(random.randint(1, 2))]
        }
    }
    
    if is_fraud and pattern:
        # Add social media red flags
        if pattern['social_media_flags']:
            metadata['social_media'] = {
                'platforms': random.sample(['facebook', 'twitter', 'instagram', 'linkedin'], 2),
                'flags': random.sample(pattern['social_media_flags'], 
                              min(2, len(pattern['social_media_flags']))),
                'last_active': fake.date_this_year().strftime('%Y-%m-%d'),
                'post_frequency': random.randint(1, 20),
                'sentiment_analysis': {
                    'positive': random.uniform(0, 0.3),
                    'negative': random.uniform(0.4, 0.9)
                }
            }
        
        # Add network analysis flags
        if pattern['network_flags']:
            metadata['network_analysis'] = {
                'connections': random.randint(5, 50),
                'fraud_connections': random.randint(1, 5),
                'flags': random.sample(pattern['network_flags'], 
                               min(2, len(pattern['network_flags']))),
                'ip_addresses': [fake.ipv4() for _ in range(random.randint(1, 3))],
                'behavioral_biometrics': {
                    'typing_speed': random.uniform(20, 80),
                    'mouse_movement': random.uniform(0.1, 0.9)
                }}
            
        
        # Add blockchain data for crypto-related fraud
        if pattern['blockchain_flags']:
            metadata['blockchain'] = {
                'wallets': [f"0x{fake.sha1()[:40]}" for _ in range(random.randint(1, 3))],
                'transactions': random.randint(1, 100),
                'flags': random.sample(pattern['blockchain_flags'], 
                              min(2, len(pattern['blockchain_flags']))),
                'token_movements': [
                    {
                        'from': f"0x{fake.sha1()[:40]}",
                        'to': f"0x{fake.sha1()[:40]}",
                        'amount': random.uniform(0.1, 50),
                        'token': random.choice(['ETH', 'BTC', 'USDT'])
                    } for _ in range(random.randint(1, 5))
                ]
            }
        
        # Add medical data for healthcare fraud
        if pattern['medical_flags']:
            metadata['medical_data'] = {
                'provider': fake.company(),
                'procedures': [
                    {
                        'code': fake.bothify(text='CPT-####'),
                        'date': fake.date_this_year().strftime('%Y-%m-%d'),
                        'cost': random.uniform(100, 5000)
                    } for _ in range(random.randint(1, 5))
                ],
                'flags': random.sample(pattern['medical_flags'], 
                              min(2, len(pattern['medical_flags'])))
            }
        
        # Add geospatial data
        if pattern['geospatial_flags']:
            metadata['location_data'] = {
                'claimed_location': {
                    'address': fake.address(),
                    'coordinates': [float(fake.latitude()), float(fake.longitude())]
                },
                'ip_locations': [
                    {
                        'ip': fake.ipv4(),
                        'country': fake.country(),
                        'vpn': random.choice([True, False])
                    } for _ in range(random.randint(1, 3))
                ],
                'flags': random.sample(pattern['geospatial_flags'], 
                              min(1, len(pattern['geospatial_flags'])))
            }
        
        # Add timing anomalies
        if random.random() > 0.7:
            metadata['timing_anomalies'] = {
                'report_delay_days': random.randint(30, 180),
                'weekend_claim': random.choice([True, False]),
                'holiday_claim': random.choice([True, False]),
                'timezone_hopping': random.choice([True, False])
            }
    
    return metadata

def generate_detailed_claim_text(is_fraud: bool, pattern: dict = None) -> str:
    """Generate realistic claim text with embedded fraud signals"""
    
    base_incidents = {
        'vehicle': ['collision', 'hit-and-run', 'theft', 'vandalism'],
        'property': ['burglary', 'fire', 'flood', 'storm damage'],
        'health': ['injury', 'illness', 'disability', 'medical emergency'],
        'cyber': ['hacking', 'data breach', 'crypto theft', 'identity theft'],
        'travel': ['trip cancellation', 'lost luggage', 'medical evacuation']
    }
    
    if is_fraud:
        incident_type = random.choice(list(base_incidents.keys()))
        
        # Create sophisticated inconsistencies
        inconsistencies = [
            ("timeline", "I was actually out of town that week but the incident occurred"),
            ("location", "The GPS data shows I was elsewhere during the claimed time"),
            ("document", "The receipts appear altered when examined closely"),
            ("witness", "The witness provided conflicting statements upon re-interview"),
            ("medical", "The treatment records don't match the claimed injuries")
        ]
        
        selected_inconsistencies = random.sample(inconsistencies, k=min(3, len(inconsistencies)))
        
        story = [
            f"On {fake.date_between(start_date='-1y', end_date='today').strftime('%B %d, %Y')}, ",
            f"I experienced {random.choice(base_incidents[incident_type])} while ",
            f"{random.choice(['traveling', 'working', 'at home', 'on vacation'])}. ",
            
            f"The incident occurred when {fake.sentence()} ",
            
            # Insert pattern-specific narrative elements
            f"{random.choice(pattern['text_keywords']).capitalize()} has complicated matters because ",
            f"{fake.sentence()}. ",
            
            # Add deliberate inconsistencies
            " ".join([
                f"Regarding the {inc[0]}, {inc[1]}. " 
                for inc in selected_inconsistencies
            ]),
            
            f"I immediately {random.choice([
                'contacted authorities',
                'took photos',
                'sought medical attention',
                'notified my insurance agent'
            ])}, though {random.choice([
                'there was some delay',
                'the response was slow',
                'documentation was incomplete'
            ])}. ",
            
            f"Supporting documents include {random.choice([
                'photographs',
                'receipts',
                'medical records',
                'police reports'
            ])}, but {random.choice([
                'some are difficult to read',
                'a few are missing dates',
                'the originals were lost'
            ])}. ",
            
            f"Additional context: {fake.paragraph(nb_sentences=3)} ",
            
            # Insert more pattern-specific content
            f"{random.choice(pattern['text_keywords']).capitalize()} further complicates this because ",
            f"{fake.sentence()}"
        ]
    else:
        incident_type = random.choice(list(base_incidents.keys()))
        
        story = [
            f"On {fake.date_between(start_date='-1y', end_date='today').strftime('%B %d, %Y')}, ",
            f"I experienced {random.choice(base_incidents[incident_type])} while ",
            f"{random.choice(['driving to work', 'at home', 'traveling abroad', 'at the office'])}. ",
            
            f"The incident occurred at precisely {fake.time(pattern='%I:%M %p')} when ",
            f"{fake.sentence()}. ",
            
            f"I immediately {random.choice([
                'called 911',
                'contacted my insurance agent',
                'documented the scene with photos',
                'sought medical attention'
            ])} and {fake.sentence()}. ",
            
            f"Independent verification includes {random.choice([
                'police report #' + fake.bothify(text='#######'),
                'security footage from ' + fake.company(),
                'witness statements from ' + fake.name(),
                'medical records from ' + fake.company()
            ])}. ",
            
            f"All documentation is complete and consistent, including ",
            f"{random.choice([
                'dated photographs',
                'itemized receipts',
                'signed witness statements',
                'official reports'
            ])}. ",
            
            f"Additional details: {fake.paragraph(nb_sentences=3)}"
        ]
    
    full_text = " ".join([s for s in story if s])
    while len(full_text.split()) < 100:  # Ensure sufficient length
        full_text += " " + fake.sentence()
    
    return full_text.replace("  ", " ").strip()

def generate_complex_claims(file_path: str, num_claims: int = 50):
    """Generate claims optimized for advanced agent detection"""
    
    claims = []
    fraud_agents = AdvancedFraudAgents()
    
    for _ in range(num_claims):
        is_fraud = random.choices([True, False], weights=[0.35, 0.65])[0]
        pattern = random.choice(list(FRAUD_PATTERNS.values())) if is_fraud else None
        
        metadata = generate_claim_metadata(is_fraud, pattern)
        
        claim_data = {
            'ClaimID': f"CLM-{str(uuid.uuid4())[:8].upper()}",
            'Claimant': fake.name(),
            'PolicyNumber': f"POL-{random.randint(1000,9999)}-{random.randint(10,99)}",
            'IncidentDate': fake.date_between(start_date='-1y', end_date='today').strftime('%Y-%m-%d'),
            'ReportDate': (
                fake.date_between(
                    start_date=datetime.now() - timedelta(days=metadata.get('timing_anomalies', {}).get('report_delay_days', 0)),
                    end_date='today'
                ).strftime('%Y-%m-%d') 
                if is_fraud 
                else fake.date_between(start_date='-30d', end_date='today').strftime('%Y-%m-%d')),
            'ClaimText': generate_detailed_claim_text(is_fraud, pattern),
            'ClaimAmount': random.randint(*pattern['amount_range']) if is_fraud else random.randint(500, 15000),
            'ClaimType': random.choice(['Vehicle', 'Property', 'Health', 'Cyber', 'Travel']),
            'FraudFlag': is_fraud,
            'FraudPattern': pattern['description'] if is_fraud else 'Legitimate',
            'Status': random.choices(
                ['Pending', 'Under Review', 'Approved', 'Rejected'],
                weights=[0.3, 0.4, 0.2, 0.1]
            )[0],
            'SupportingDocs': ", ".join([
                f"{fake.word()}_{random.choice(['report', 'photos', 'statement'])}_{random.randint(1,100)}.pdf"
                for _ in range(random.randint(2, 5))
            ]),
            'Metadata': json.dumps(metadata, indent=2)
        }
        
        # Create explicit connections for network analysis
        if is_fraud and random.random() > 0.6:
            related_claims = [f"CLM-{str(uuid.uuid4())[:8].upper()}" for _ in range(random.randint(1, 3))]
            metadata['related_claims'] = related_claims
            claim_data['RelatedClaims'] = ", ".join(related_claims)
        
        # Perform agent analysis (for demonstration, we'll simulate this)
        if is_fraud:
            metadata['agent_analysis'] = {
                "simulated_analysis": True,
                "potential_fraud_indicators": [
                    {"type": k, "score": random.uniform(0.6, 0.95)} 
                    for k in ['social', 'network', 'behavioral']
                ]
            }
        
        claims.append(claim_data)
    
    df = pd.DataFrame(claims)
    
    # Calculate derived fields
    df['DaysToReport'] = (
        pd.to_datetime(df['ReportDate']) - 
        pd.to_datetime(df['IncidentDate'])
    ).dt.days
    df['AmountPerDay'] = df['ClaimAmount'] / df['DaysToReport'].clip(lower=1)
    
    # Save to Excel with enhanced formatting
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Main claims sheet
        df.to_excel(writer, sheet_name='Claims', index=False)
        
        # Create hidden metadata sheet for agents
        metadata_df = pd.DataFrame([json.loads(m) for m in df['Metadata']])
        metadata_df.insert(0, 'ClaimID', df['ClaimID'])
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Formatting
        workbook = writer.book
        
        # Hide metadata sheet
        workbook['Metadata'].sheet_state = 'hidden'
        
        # Format main sheet
        claims_sheet = writer.sheets['Claims']
        
        # Set column widths
        col_widths = {
            'A': 15, 'B': 20, 'C': 18, 'D': 12, 'E': 12,
            'F': 80, 'G': 15, 'H': 15, 'I': 10, 'J': 25,
            'K': 15, 'L': 30, 'M': 50, 'N': 12, 'O': 12
        }
        for col, width in col_widths.items():
            claims_sheet.column_dimensions[col].width = width
        
        # Add conditional formatting for fraud
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        for row in claims_sheet.iter_rows(min_row=2, max_row=claims_sheet.max_row, min_col=9, max_col=9):
            for cell in row:
                if cell.value == True:
                    cell.fill = red_fill
                else:
                    cell.fill = green_fill
        
        # Wrap text for claim text and metadata
        for col in ['F', 'M']:
            for row in claims_sheet.iter_rows(min_row=2, max_row=claims_sheet.max_row, min_col=ord(col)-64, max_col=ord(col)-64):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    print(f"\nGenerated {num_claims} advanced claims at: {file_path}")
    print(f"Fraudulent claims: {sum(df['FraudFlag'])}")
    print(f"Fraud patterns: {df[df['FraudFlag']]['FraudPattern'].value_counts().to_dict()}")
    
    return df

# Generate the enhanced claims file with agent analysis capabilities
if __name__ == "__main__":
    claims_data = generate_complex_claims("advanced_fraud_claims_with_agents.xlsx", num_claims=2)