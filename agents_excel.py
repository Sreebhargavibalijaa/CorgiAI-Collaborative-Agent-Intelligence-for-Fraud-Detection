import pandas as pd
import numpy as np
import json
import logging
import hashlib
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from autogen import AssistantAgent, UserProxyAgent, GroupChat, GroupChatManager
import openpyxl
from openpyxl.styles import Font, PatternFill
import threading
import os
from dotenv import load_dotenv

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

class Config:
    FRAUD_THRESHOLD = 0.7
    CONSENSUS_THRESHOLD = 3
    HISTORY_FILE = "fraud_history.json"
    OUTPUT_EXCEL_FORMAT = {
        'font': {'name': 'Arial', 'size': 11},
        'header': {'bold': True, 'bg_color': '#4472C4', 'font_color': 'white'}
    }
    LLM_CONFIG = {
        "config_list": [
            {
                "model": "gpt-3.5-turbo",
                "api_key": os.getenv("OPENAI_API_KEY")
            }
        ],
        "temperature": 0.3,
        "timeout": 60
    }
    GPT4_CONFIG = {
        "config_list": [
            {
                "model": "gpt-4",
                "api_key": os.getenv("OPENAI_API_KEY")
            }
        ],
        "temperature": 0.3,
        "timeout": 60
    }



class MCPIntegration:
    def __init__(self):
        self.scores = {}
        self.history = []
        self.lock = threading.Lock()
        logger.info("MCP initialized")

    def update_score(self, claim_id: str, agent: str, score: float, metadata: Dict = None) -> bool:
        try:
            if not 0 <= score <= 1:
                raise ValueError(f"Invalid score {score}. Must be between 0-1")
                
            timestamp = datetime.now().isoformat()
            
            with self.lock:
                if claim_id not in self.scores:
                    self.scores[claim_id] = {}
                
                self.scores[claim_id][agent] = {
                    'score': score,
                    'timestamp': timestamp,
                    'metadata': metadata or {}
                }
                
                self.history.append({
                    'claim_id': claim_id,
                    'agent': agent,
                    'score': score,
                    'timestamp': timestamp,
                    'metadata': metadata
                })
                
            return True
            
        except Exception as e:
            logger.error(f"Score update failed: {e}")
            return False

    def get_consensus(self, claim_id: str) -> Tuple[bool, Dict]:
        try:
            with self.lock:
                if claim_id not in self.scores:
                    return False, {}
                
                scores_data = self.scores[claim_id]
                simple_scores = {k: v['score'] for k, v in scores_data.items()}
                
                fraud_count = sum(1 for s in simple_scores.values() 
                                if s >= Config.FRAUD_THRESHOLD)
                has_consensus = fraud_count >= Config.CONSENSUS_THRESHOLD
                
                avg_score = sum(simple_scores.values()) / len(simple_scores) if simple_scores else 0
                max_score = max(simple_scores.values()) if simple_scores else 0
                
                return has_consensus, {
                    'scores': simple_scores,
                    'average_score': round(avg_score, 2),
                    'max_score': round(max_score, 2),
                    'fraud_count': fraud_count,
                    'agent_details': scores_data
                }
                
        except Exception as e:
            logger.error(f"Consensus check failed: {e}")
            return False, {}

class ExcelProcessor:
    REQUIRED_COLUMNS = ['ClaimID', 'Claimant', 'ClaimText']
    OPTIONAL_COLUMNS = [
        'DateOfIncident', 
        'ClaimAmount', 
        'PolicyNumber', 
        'ContactEmail', 
        'SupportingDocs',
        'MedicalCodes',
        'Location',
        'TransactionHashes'
    ]
    
    @staticmethod
    def generate_template(output_path: str = "claims_template.xlsx"):
        sample_data = {
            'ClaimID': ['CLM-001', 'CLM-002', 'CLM-003'],
            'Claimant': ['JohnDoe', 'JaneSmith', 'MikeBrown'],
            'ClaimText': ['Car accident on Main St', 'Stolen laptop', 'Skiing injury'],
            'DateOfIncident': ['2023-01-15', '2023-02-20', '2023-03-05'],
            'ClaimAmount': [2500, 1800, 3500],
            'PolicyNumber': ['POL-1001', 'POL-1002', 'POL-1003'],
            'ContactEmail': ['john@test.com', 'jane@test.com', 'mike@test.com'],
            'SupportingDocs': ['police_report.pdf', 'receipt.jpg', 'medical_report.pdf'],
            'MedicalCodes': ['', 'ICD-10:S72.8X1A', 'CPT:99214'],
            'Location': ['40.7128,-74.0060', '34.0522,-118.2437', '39.7392,-104.9903'],
            'TransactionHashes': ['', '0xabc...123', '']
        }
        
        df = pd.DataFrame(sample_data)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            for col in worksheet.iter_cols(1, len(df.columns)):
                col[0].font = Font(bold=True, color="FFFFFF")
                col[0].fill = PatternFill("solid", fgColor="4472C4")
        
        logger.info(f"Template generated: {output_path}")
        return output_path

    @staticmethod
    def validate_input(file_path: str) -> bool:
        try:
            df = pd.read_excel(file_path)
            missing = [col for col in ExcelProcessor.REQUIRED_COLUMNS if col not in df.columns]
            if missing:
                raise ValueError(f"Missing required columns: {missing}")
            return True
        except Exception as e:
            logger.error(f"Excel validation failed: {e}")
            return False

    @staticmethod
    def save_results(results: List[Dict], input_path: str) -> str:
        output_path = input_path.replace('.xlsx', '_processed.xlsx')
        
        # Flatten the agent_scores structure for Excel output
        processed_results = []
        for result in results:
            flat_result = {
                'ClaimID': result.get('claim_id'),
                'Claimant': result.get('claimant'),
                'OriginalClaim': result.get('claim_text'),
                'Decision': result.get('decision'),
                'ProcessingTime': result.get('decision_timestamp'),
                'ProcessingDurationSeconds': result.get('processing_time_seconds'),
                'AverageScore': result.get('agent_scores', {}).get('average_score', 0),
                'MaxScore': result.get('agent_scores', {}).get('max_score', 0),
                'FraudVotes': result.get('agent_scores', {}).get('fraud_count', 0),
                **result.get('metadata', {})
            }
            
            # Add individual agent scores
            for agent, score in result.get('agent_scores', {}).get('scores', {}).items():
                flat_result[f'{agent}_score'] = score
            
            processed_results.append(flat_result)
        
        df = pd.DataFrame(processed_results)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            worksheet = writer.sheets['Results']
            
            # Format headers
            for col in worksheet.iter_cols(1, len(df.columns)):
                col[0].font = Font(bold=True, color="FFFFFF")
                col[0].fill = PatternFill("solid", fgColor="4472C4")
                
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        logger.info(f"Results saved to: {output_path}")
        return output_path

class CorgiAgentSystem:
    def __init__(self):
        self.mcp = MCPIntegration()
        self.agents = self._initialize_agents()
        self.agent_versions = {
            "social": "2.1",
            "network": "2.1",
            "blockchain": "1.3",
            "medical": "1.0",
            "geospatial": "1.2",
            "decision": "3.0"
        }
        logger.info("Advanced Corgi Agent System initialized")

    def _initialize_agents(self) -> Dict[str, AssistantAgent]:
        """Initialize the complete set of advanced analysis agents"""
        agents = {
            "social": AssistantAgent(
                name="SocialAnalystPro",
                system_message="""Analyze social media for fraud indicators:
                - Post/claim timeline analysis
                - Geolocation verification
                - Sentiment inconsistency
                - Image metadata examination
                Output JSON: {
                    "score":0-1, 
                    "confidence":0-1,
                    "timeline_analysis": {...},
                    "geolocation_verification": bool,
                    "media_analysis": [...]
                }""",
                llm_config=Config.LLM_CONFIG
            ),
            "network": AssistantAgent(
                name="NetworkThreatIntel",
                system_message="""Examine network patterns for anomalies:
                - IP reputation analysis
                - Device fingerprinting
                - Behavioral biometrics
                - Temporal patterns
                Output JSON: {
                    "score":0-1,
                    "threat_level":"low/medium/high",
                    "suspicious_ips": [...],
                    "behavioral_anomalies": [...]
                }""",
                llm_config=Config.LLM_CONFIG
            ),
            "blockchain": AssistantAgent(
                name="BlockchainForensics",
                system_message="""Analyze blockchain transactions:
                - Wallet transaction graph
                - Smart contract interactions
                - Mixer/tumbler detection
                - Dark web affiliations
                Output JSON: {
                    "score":0-1,
                    "wallet_risk":0-1,
                    "suspicious_transactions": [...],
                    "mixer_usage": bool
                }""",
                llm_config=Config.LLM_CONFIG
            ),
            "medical": AssistantAgent(
                name="MedicalClaimsExpert",
                system_message="""Evaluate medical claims:
                - Procedure-code consistency
                - Treatment duration analysis
                - Provider reputation
                - Historical claim patterns
                Output JSON: {
                    "score":0-1,
                    "procedure_consistency":0-1,
                    "treatment_anomalies": [...],
                    "provider_risk":0-1
                }""",
                llm_config=Config.LLM_CONFIG
            ),
            "geospatial": AssistantAgent(
                name="GeospatialAnalyst",
                system_message="""Verify geospatial claims:
                - Location verification
                - Weather correlation
                - Traffic/accident data
                - Satellite imagery analysis
                Output JSON: {
                    "score":0-1,
                    "location_verified": bool,
                    "weather_consistency":0-1,
                    "timestamp_analysis": {...}
                }""",
                llm_config=Config.LLM_CONFIG
            ),
            "decision": AssistantAgent(
                name="DecisionEnginePro",
                system_message="""Make final claim determination:
                1. Aggregate agent findings
                2. Apply business rules
                3. Consider regulatory requirements
                4. Evaluate fraud probability
                Final JSON output: {
                    "decision":"APPROVE/REJECT/ESCALATE",
                    "confidence":0-1,
                    "risk_factors": [...],
                    "recommendations": [...],
                    "audit_required": bool
                }""",
                llm_config=Config.GPT4_CONFIG
            )
        }

        def make_advanced_handler(agent_name: str):
            def handler(sender, recipient, context):
                try:
                    if isinstance(context, dict):
                        self.mcp.update_score(
                            context.get('claim_id'),
                            agent_name,
                            float(context.get('score', 0)),
                            metadata={
                                "analysis": context,
                                "timestamp": datetime.now().isoformat()
                            }
                        )
                except Exception as e:
                    logger.error(f"Handler error for {agent_name}: {e}")
                return False
            return handler

        for name, agent in agents.items():
            if name != "decision":
                agent.register_reply(
                    trigger=AssistantAgent,
                    reply_func=make_advanced_handler(name),
                    position=1,
                    config={"priority": 1}
                )
        
        return agents

    def process_claim(self, claim_text: str, claimant: str, metadata: Dict = None) -> Dict:
        """Process a claim through the complete agent network"""
        metadata = metadata or {}
        claim_id = metadata.get('ClaimID', hashlib.sha256(f"{claimant}{datetime.now().timestamp()}".encode()).hexdigest())
        
        try:
            groupchat = GroupChat(
                agents=list(self.agents.values()),
                messages=[],
                max_round=10,
                speaker_selection_method="auto"
            )
            
            manager = GroupChatManager(
                groupchat=groupchat,
                llm_config=Config.LLM_CONFIG
            )
            
            user_proxy = UserProxyAgent(
                name="ClaimProcessor",
                human_input_mode="TERMINATE",
                code_execution_config=False,
                default_auto_reply="Continue..."
            )
            
            processing_start = datetime.now()
            user_proxy.initiate_chat(
                manager,
                message=json.dumps({
                    "action": "process_claim",
                    "claim_id": claim_id,
                    "claim_text": claim_text,
                    "claimant": claimant,
                    "metadata": metadata
                }, indent=2)
            )
            
            is_fraud, agent_scores = self.mcp.get_consensus(claim_id)
            
            # Generate comprehensive report
            report = {
                "claim_id": claim_id,
                "claimant": claimant,
                "claim_text": claim_text,
                "decision": "REJECT" if is_fraud else "APPROVE",
                "decision_timestamp": datetime.now().isoformat(),
                "processing_time_seconds": round((datetime.now() - processing_start).total_seconds(), 2),
                "agent_scores": agent_scores,
                "agent_versions": self.agent_versions,
                "metadata": metadata
            }
            
            return report
            
        except Exception as e:
            logger.error(f"Failed to process claim {claim_id}: {e}")
            return {
                "claim_id": claim_id,
                "claimant": claimant,
                "claim_text": claim_text,
                "decision": "ERROR",
                "decision_timestamp": datetime.now().isoformat(),
                "error": str(e),
                "metadata": metadata
            }

    def process_excel(self, file_path: str) -> str:
        """Batch process claims from Excel file"""
        try:
            if not ExcelProcessor.validate_input(file_path):
                raise ValueError("Invalid Excel file structure")
            
            df = pd.read_excel(file_path)
            results = []
            
            for _, row in df.iterrows():
                metadata = {col: row[col] for col in df.columns if pd.notna(row[col])}
                result = self.process_claim(
                    claim_text=row['ClaimText'],
                    claimant=row['Claimant'],
                    metadata=metadata
                )
                results.append(result)
            
            output_path = ExcelProcessor.save_results(results, file_path)
            logger.info(f"Processed {len(results)} claims successfully")
            return output_path
            
        except Exception as e:
            logger.error(f"Excel processing failed: {e}")
            raise

def main():
    try:
        system = CorgiAgentSystem()
        
        template_path = "claims_template.xlsx"
        if not os.path.exists(template_path):
            ExcelProcessor.generate_template(template_path)
            print(f"\nTemplate created: {template_path}")
            print("Please fill it with your claims and run again.")
            return
        
        input_path = "/Users/sreebhargavibalija/Desktop/corgi/corgi-fraud/advanced_fraud_claims_with_agents.xlsx"
        # if not os.path.exists(input_path):
        #     input_path = template_path
        
        result_file = system.process_excel(input_path)
        print(f"\nProcessing complete. Results saved to:\n{result_file}")
        
    except Exception as e:
        logger.critical(f"System error: {e}")
        raise

if __name__ == "__main__":
    main()